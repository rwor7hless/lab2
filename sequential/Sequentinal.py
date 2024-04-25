import string

import openpyxl
import pickle
import numpy as np
#import matplotlib.pyplot as plt
from keras.src.optimizers import Adam
from tensorflow.keras.models import Sequential
from tensorflow.keras.layers import Embedding, Flatten, Dense
from tensorflow.keras.preprocessing.text import Tokenizer
from tensorflow.keras.preprocessing.sequence import pad_sequences
from sklearn.model_selection import train_test_split

# Загрузка списка слов из файла со словами
def load_word_list(file_path):
    with open(file_path, 'r', encoding='utf-8') as file:
        word_list = file.read().splitlines()
    return word_list

# Предварительная обработка текстов, оставляя только слова из списка
# def filter_text(text, word_list):
#     words = text.split()
#     filtered_words = [word for word in words if any(root in word for root in word_list)]
#     filtered_text = ' '.join(filtered_words)
#     return filtered_text


def remove_punctuation(text):
    # Удаляем знаки препинания из текста
    cleaned_text = ''.join(char for char in text if char not in string.punctuation).lower()
    return cleaned_text
def filter_text(text, word_list):
    text = remove_punctuation(text)
    words = text.split()
    filtered_text = []
    negate = False  # флаг, который будет указывать на текущий отрицательный контекст

    for word in words:
        if word == "не":  # если слово "не" обнаружено, меняем флаг отрицания
            negate = True
        elif any(root in word for root in word_list):  # проверяем наличие корня из списка слов в текущем слове
            if negate:  # если флаг отрицания установлен
                filtered_text.append("не " + word)
                negate = False
            else:  # если отрицание неактивно, добавляем слово без "не"
                filtered_text.append(word)
        elif "не" in word:  # если слово содержит "не", но не в списке отрицательных слов
            filtered_text.append(word)
    print(filtered_text)
    return ' '.join(filtered_text)
# Загрузка данных из Excel файла
def load_data(file_path, word_list):
    workbook = openpyxl.load_workbook(file_path, read_only=True)
    sheet = workbook.active

    texts = []
    labels = []

    for row in sheet.iter_rows(min_row=1, max_row=1000, values_only=True):
        text = row[0] if row[0] is not None else ""
        filtered_text = filter_text(text, word_list)
        texts.append(filtered_text)

        label = row[1] if row[1] is not None else 0
        labels.append(int(label))

    print(texts, labels)
    return texts, labels

# Путь к файлу с корнями слов
word_list_path = 'words.txt'
word_list = load_word_list(word_list_path)

# Путь к Excel файлу с данными
excel_file_path = '../reviews.xlsx'

# Загрузка данных
texts, labels = load_data(excel_file_path, word_list)

# Предварительная обработка обучающих текстов
filtered_texts = [filter_text(text, word_list) for text in texts]

# Токенизация текстов и преобразование их в последовательности
tokenizer = Tokenizer(num_words=len(word_list), oov_token='<OOV>')
tokenizer.fit_on_texts(texts)
sequences = tokenizer.texts_to_sequences(filtered_texts)

# Паддинг последовательностей
padded_sequences = pad_sequences(sequences, maxlen=500, padding='post', truncating='post')

# Разделение данных на обучающий и тестовый наборы
X_train, X_test, y_train, y_test = train_test_split(padded_sequences, labels, shuffle=False, test_size=0.001, random_state=42)

# Создание модели
model = Sequential()
model.add(Embedding(input_dim=len(word_list), output_dim=2, input_length=500))
model.add(Flatten())
model.add(Dense(16, activation='relu'))
#model.add(Dropout(0.5))
model.add(Dense(8, activation='relu'))
model.add(Dense(1, activation='sigmoid'))

#C этими параметрами оutput_dim где-то 50-100 ставить стоит
# model.add(Dense(305, activation='leaky_relu'))
# model.add(Dropout(0.5))
# model.add(Dense(150, activation='relu'))
# model.add(Dense(75, activation='softmax'))
# model.add(Dense(1, activation='sigmoid'))


# Компиляция модели
model.compile(optimizer=Adam(lr=0.0001), loss='binary_crossentropy', metrics=['accuracy'])

# Обучение модели
history = model.fit(X_train, np.array(y_train), epochs=100, batch_size=16, validation_split=0.001)

# Предсказание на обучающих данных
predictions_train = model.predict(X_train)

# Вывод номера отзыва и предполагаемой тональности на обучающих данных
for i in range(len(predictions_train)):
    review_number = i + 1
    predicted_sentiment = "Положительный" if predictions_train[i] > 0.5 else "Отрицательный"
    print(f"Отзыв #{review_number}: Предполагаемая тональность - {predicted_sentiment}")


# Загрузка тестовых данных из Excel файла
def load_test_data(file_path):
    workbook = openpyxl.load_workbook(file_path, read_only=True)
    sheet = workbook.active

    test_texts = []

    for row in sheet.iter_rows(min_row=1, max_row=100, values_only=True):  # Указать правильное количество строк
        text = row[0] if row[0] is not None else ""
        test_texts.append(text)

    return test_texts

# Путь к вашему Excel файлу с тестовыми данными
test_excel_file_path = '../tests/50pos/50pos.xlsx'

# Загрузка тестовых данных
test_texts = load_test_data(test_excel_file_path)

# Предварительная обработка тестовых текстов
filtered_test_texts = [filter_text(text, word_list) for text in test_texts]

# Токенизация тестовых текстов и преобразование их в последовательности
test_sequences = tokenizer.texts_to_sequences(filtered_test_texts)

# Паддинг тестовых последовательностей
padded_test_sequences = pad_sequences(test_sequences, maxlen=500, padding='post', truncating='post')

# Предсказание на тестовых данных
predictions_test = model.predict(padded_test_sequences)

# Вывод номера отзыва и предполагаемой тональности на тестовых данных
for i in range(len(predictions_test)):
    review_number = i + 1
    predicted_sentiment = "Положительный" if predictions_test[i] > 0.5 else "Отрицательный"
    print(f"Отзыв #{review_number}: Предполагаемая тональность - {predicted_sentiment}")


# Предсказание на тестовых данных
predictions_test = model.predict(padded_test_sequences)

# Порог для классификации
threshold = 0.5

# Подсчет положительных и отрицательных предсказаний
positive_predictions = np.sum(predictions_test > threshold)
negative_predictions = len(predictions_test) - positive_predictions

# Вывод результатов
print(f"Общее количество положительных предсказаний: {positive_predictions} / всего 36")
print(f"Общее количество отрицательных предсказаний: {negative_predictions} / всего 64")


# # Построение графика точности
# plt.plot(history.history['accuracy'])
# plt.plot(history.history['val_accuracy'])
# plt.title('Model Accuracy')
# plt.xlabel('Epoch')
# plt.ylabel('Accuracy')
# plt.legend(['Train', 'Validation'], loc='upper left')
# plt.show()
#
# # Построение графика функции потерь
# plt.plot(history.history['loss'])
# plt.plot(history.history['val_loss'])
# plt.title('Model Loss')
# plt.xlabel('Epoch')
# plt.ylabel('Loss')
# plt.legend(['Train', 'Validation'], loc='upper left')
# plt.show()


# print(len(word_list))


model.save('model_seq.h5')
with open('tokenizer.pickle', 'wb') as handle:
    pickle.dump(tokenizer, handle, protocol=pickle.HIGHEST_PROTOCOL)
# print("Длина данных перед токенизацией:", len(texts))
# print("Длина обучающих данных после токенизации:", len(X_train))
# print("Длина предсказаний для обучающих данных:", len(predictions_train))
