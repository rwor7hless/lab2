import string
import openpyxl

weightsReview = [175, 6, -11, 232, -379, 66, 7, 7, 619, 258, 12, -80, -139, -932, 1, -82, 7, 2, 1001, 44, 8, 58, 2, -90, 54, 5, 4, -86, 7, -6, 4, -10, -97, 53, 4, 84, 10, 3, 1, -63, 5, 3, 312, 1, 14, 9, 5, 2048, 2, 139, -289, 7, 8, -479, 5, 6, 43, 164, -48, 875, 63, 6, 1188, 7, -338, 7, 9, -75, 17, -337, 9, 4, 7, 419, 753, 15, -243, 8, -106, -275, -39, -17, 2, 5, 2, 18, -690, 9, -167, 7, 4, 3, 8, 5, 8, -45, 2, 4, 6, 1, 4, 9, 15, 12, 9, -322, -71, 8, 9, 1, 9, 7, 1, -174, -87, -89, 841, 9, 339, 744, 32, -432, -3, 3, 1, 1, 643, 408, 11, 244, 501, 7, 1001, 9, 17, 1001, -44, 99, -80, -568, -365, 5, 366, 872, 7, -745, 15, -147, -1047, 25, 168, 127, -375, -10, 4, 1001, 126, 1001, 594, -24, 131, 3, 284, 547, 242, 612, 42, -53, -188, 460, 45, 1366, 483, 117, 1, -125, 4, 902, -30, 9, 82, -476, 873, 5, 31, -327, 4, 8, 148, 4, 1001, 86, -335, 28, 1000, 11, 337, 8, -96, -78, 7, -127, -384, 2, 390, 744, 31, 312, 132, 6, -385, -186, 13, 5, 53, 589, -50, 9, -974, 12, 8, 19, -259, 5, 197, 1, 12, 1011, 4, 232, 13, 105, 61, -47, 1003, 253, 2, 218, 7, -668, 171, 0, 14, 6, 4, 70, -902, 4, 245, 3, 230, 233, 1001, 4, 3, 9, 7, 9, 7, 1025, 167, 9, 310, 51, -97, 8, -2, 7, 172, -53, 138, 102, 34, 1, 503, 2, 22, 1, 139, 129, 424, 259, 8, -155, 292, 3, -310, -589, 1015, 593, -309, -82, 1289, 641, 49, 923, 182, 6, 123, 6, 21, 0, 13, 36, 3, 558, 161, 359, -66, 18]

def import_words(filepath):
    import os
    if not os.path.exists(filepath):
        print(f"Файл {filepath} не найден.")
    else:
        with open(filepath, 'r', encoding="utf-8") as file:
            unique_word = file.read()
            words = unique_word.splitlines()
            return words


def for_front(review):
    print(f"Введенный отзыв: {review}")
    review = ''.join(char for char in review if char not in string.punctuation).lower()
    review = review.split()
    merged_review = []
    i = 0
    while i < len(review):
        if review[i] == "не" and i+1 < len(review):
            merged_review.append(review[i] + " " +  review[i+1])
            i += 2
        else:
            merged_review.append(review[i])
            i += 1
    print(merged_review)

    len_dataset, len_review = len(dataset_words), len(merged_review)
    binaryreview = [0] * len_dataset
    print(review)
    print()

    for word in merged_review:
        for i in range(len_dataset):
            if dataset_words[i] in word:
                binaryreview[i] = 1
                print(dataset_words[i])
                break  # прерываем цикл, если корень найден


    print(binaryreview)
    print(len(weightsReview), " ", len_review, " ", len_dataset)
    for oneReview in range(len_review):
        reviewSum = 0
        for uniqWord in range(len_dataset):
            reviewSum += weightsReview[uniqWord] * binaryreview[uniqWord]
            if binaryreview[uniqWord] == 1:
                print(dataset_words[uniqWord], ' ', weightsReview[uniqWord])
        print(reviewSum)
        if reviewSum > 1000:
            return(1)
        else:
            return(0)


filepath = "../Neuron/unique_words.txt"
dataset_words = import_words(filepath)
print(dataset_words)

# //////////////////////////////////////////////////////////////////////////////////////////////////////////////////


def impoport(destination):
    # Открываем файл Excel
    workbook = openpyxl.load_workbook(destination)

    # Выбираем активный лист (первый лист в книге)
    sheet = workbook.active

    # Создаем список для хранения текста из каждой ячейки столбца
    cell_text_list = []

    # Проходим по каждой ячейке в столбце с отзывами
    for cell in sheet['A']:
        if cell.value is not None:
            print(cell.value)
            cell_text_list.append(str(cell.value))

    # Закрываем файл Excel
    workbook.close()
    # Убираем возможные дубликаты
    return cell_text_list


def remove_punctuation(text):
    # Удаляем знаки препинания из текста, оставляя пробелы
    cleaned_text = ''.join(char if char not in string.punctuation else ' ' for char in text)
    return cleaned_text


def neuron_test_excel(destination):

    # Открываем файл Excel
    workbook = openpyxl.load_workbook(destination)

    # Выбираем активный лист (первый лист в книге)
    sheet = workbook.active

    # Массивы с отзывами и словами для датасета
    all_revs = impoport(destination)
    print(all_revs)
    with open(filepath, 'r', encoding="utf-8") as file:
        dataset_words = [line.strip() for line in file.readlines()]

    # Проходим по каждому отзыву и анализируем его на наличие слов из датасета
    for row, review in enumerate(all_revs, start=2):
        clear_review = remove_punctuation(review).lower().split()
        binary_review = [0] * len(dataset_words)
        for word in clear_review:
            for i, dataset_word in enumerate(dataset_words):
                if dataset_word in word:
                    binary_review[i] = 1
                    break

        review_sum = sum(weight * binary for weight, binary in zip(weightsReview, binary_review))
        if review_sum > 1000:
            sentiment = 1
        else:
            sentiment = 0

        sheet.cell(row=row, column=2, value=sentiment)
    # Сохраняем книгу Excel
    workbook.save(destination)